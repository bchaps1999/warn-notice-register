"""Massachusetts — ported from upstream warn-scraper PR #787 (author riordan).

Changes from the PR: mass.gov's Akamai protection rejects browser-UA
mismatches but (currently) accepts plain requests with its default UA, so we
try that first and fall back to the Zyte proxy only if blocked and a
ZYTE_API_KEY is configured. Everything else (workbook + weekly CSV parsing)
follows the PR.
"""

from __future__ import annotations

import csv
import logging
import os
from datetime import datetime
from pathlib import Path
from urllib.parse import unquote, urljoin

import niquests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from warn import utils
from warn.cache import Cache

logger = logging.getLogger(__name__)

BASE_URL = "https://www.mass.gov"
INDEX_URL = (
    "https://www.mass.gov/info-details/worker-adjustment-and-retraining-"
    "notification-act-warn-layoff-and-closure-updates"
)

CANONICAL_HEADER = [
    "RECEIVED",
    "EMPLOYER",
    "CITY/TOWN",
    "REGION",
    "DATE(S) OF LAYOFFS",
    "# EMPLOYEES IMPACTED",
]

# Older fiscal-year workbooks (FY2022, FY2023) use one worksheet per region
# with this fixed column order; region comes from the worksheet name.
REGION_SHEET_ORDER = [
    "RECEIVED",
    "EMPLOYER",
    "CITY/TOWN",
    "DATE(S) OF LAYOFFS",
    "# EMPLOYEES IMPACTED",
]


# Akamai rejects on UA/TLS-fingerprint mismatch: requests' TLS gets 403,
# while niquests (utls) with its own default UA passes. Do NOT send a
# browser User-Agent — that reintroduces the mismatch. File downloads
# additionally require the session cookies from the index page + a Referer.
_session = niquests.Session()


def _get(url: str) -> bytes:
    r = _session.get(url, headers={"Referer": INDEX_URL}, timeout=120)
    if r.status_code == 403 and os.environ.get("ZYTE_API_KEY"):
        logger.debug("MA: 403 from %s, retrying via Zyte", url)
        raw, _ = utils.get_with_zyte(url)
        return raw
    r.raise_for_status()
    return r.content


def scrape(
    data_dir: Path = utils.WARN_DATA_DIR,
    cache_dir: Path = utils.WARN_CACHE_DIR,
) -> Path:
    cache = Cache(cache_dir)
    html = _get(INDEX_URL).decode("utf-8", errors="replace")
    cache.write("ma/index.html", html)

    soup = BeautifulSoup(html, "html.parser")
    excel_urls, csv_urls = _find_source_links(soup)
    logger.debug("MA: %d workbook(s), %d CSV(s)", len(excel_urls), len(csv_urls))

    master_list: list = []
    for url in excel_urls:
        name = url.split("/doc/")[1].split("/")[0]
        excel_path = cache.write_binary(f"ma/{name}.xlsx", _get(url))
        master_list.extend(_parse_workbook(excel_path))

    for url in csv_urls:
        name = os.path.basename(unquote(url))
        raw = _get(url)
        cache.write_binary(f"ma/{name}", raw)
        rows = list(csv.reader(raw.decode("utf-8", errors="replace").splitlines()))
        master_list.extend(_parse_flat_rows(rows))

    deduped = _dedupe(master_list)
    logger.debug("MA: %d rows, %d after dedupe", len(master_list), len(deduped))

    data_path = data_dir / "ma.csv"
    utils.write_dict_rows_to_csv(
        data_path, CANONICAL_HEADER, deduped, extrasaction="raise"
    )
    return data_path


def _find_source_links(soup: BeautifulSoup) -> tuple:
    excel_urls: set = set()
    csv_urls: set = set()
    for link in soup.find_all("a", href=True):
        href = link["href"]
        lower = href.lower()
        if "/doc/" in lower and "warn-report" in lower:
            excel_urls.add(urljoin(BASE_URL, href))
        elif lower.endswith(".csv") and "warn" in lower:
            csv_urls.add(urljoin(BASE_URL, href))
    return sorted(excel_urls), sorted(csv_urls)


def _parse_workbook(excel_path: Path) -> list:
    workbook = load_workbook(filename=excel_path, read_only=True, data_only=True)
    rows: list = []
    for sheet in workbook.worksheets:
        sheet_rows = [list(r) for r in sheet.iter_rows(values_only=True)]
        if _has_region_column(sheet_rows):
            rows.extend(_parse_flat_rows(sheet_rows))
        else:
            rows.extend(_parse_region_sheet(sheet_rows, sheet.title.strip()))
    workbook.close()
    return rows


def _has_region_column(rows: list) -> bool:
    header_index = _find_header_row(rows)
    if header_index is None:
        return False
    return any(_norm(cell) == "REGION" for cell in rows[header_index])


def _find_header_row(rows: list):
    for i, row in enumerate(rows):
        labels = {_norm(cell) for cell in row}
        if "EMPLOYER" in labels or "COMPANY NAME" in labels:
            return i
    return None


def _parse_flat_rows(rows: list) -> list:
    header_index = _find_header_row(rows)
    if header_index is None:
        return []
    positions: dict = {}
    for col, cell in enumerate(rows[header_index]):
        label = _norm(cell)
        if label in CANONICAL_HEADER and label not in positions:
            positions[label] = col
    parsed = []
    for row in rows[header_index + 1 :]:
        record = {
            field: _clean(row[pos]) if pos < len(row) else ""
            for field, pos in positions.items()
        }
        record = {field: record.get(field, "") for field in CANONICAL_HEADER}
        if _keep_row(record):
            parsed.append(record)
    return parsed


def _parse_region_sheet(rows: list, region: str) -> list:
    header_index = _find_header_row(rows)
    if header_index is None:
        return []
    parsed = []
    for row in rows[header_index + 1 :]:
        record = {field: "" for field in CANONICAL_HEADER}
        for pos, field in enumerate(REGION_SHEET_ORDER):
            if pos < len(row):
                record[field] = _clean(row[pos])
        record["REGION"] = region
        if _keep_row(record):
            parsed.append(record)
    return parsed


def _keep_row(record: dict) -> bool:
    employer = record.get("EMPLOYER", "")
    if not employer:
        return False
    if _norm(employer) in ("EMPLOYER", "COMPANY NAME"):
        return False
    if _norm(employer).startswith("TOTAL"):
        return False
    return True


def _dedupe(rows: list) -> list:
    seen = set()
    deduped = []
    for row in rows:
        key = tuple(row[field] for field in CANONICAL_HEADER)
        if key not in seen:
            seen.add(key)
            deduped.append(row)
    return deduped


def _clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return f"{value.month}/{value.day}/{value.year}"
    return str(value).strip()


def _norm(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split()).upper()

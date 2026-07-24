"""Texas — patched from upstream warn-scraper tx.py.

twc.texas.gov sits behind AWS WAF (CloudFront), which serves GitHub runner
IPs a JavaScript challenge (HTTP 202, window.awsWafCookie) on every path —
index page and xlsx assets alike — so the upstream scraper fails with
"Scraper isn't scraping" on CI (7 straight days as of 2026-07-23; upstream
#767 blamed Cloudflare, but the probe workflow showed AWS WAF).

Strategy: try plain HTTP first (fine from residential IPs). When discovery
yields no links, load the index page in headless Chrome — the browser
passes the WAF challenge automatically — then copy the aws-waf-token
cookies and user agent into the HTTP session for the workbook downloads.
Downloads are validated as real xlsx (zip magic) so a challenge page can
never masquerade as a workbook.
"""

from __future__ import annotations

import logging
import random
import re
from datetime import date
from pathlib import Path
from time import sleep, time

import niquests as requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from warn import utils
from warn.cache import Cache

logger = logging.getLogger(__name__)

INDEX_URL = "https://www.twc.texas.gov/data-reports/warn-notice"
ASSET_ROOT = "https://www.twc.texas.gov"
HREF_PATTERN = re.compile(r"^/sites/default/files/oei/docs/warn-act-listings-")
# Years covered by the yearly workbooks; earlier years come from the BLN
# historical file. Upstream keeps 2019+, though the index lists 2020+ now.
FIRST_YEAR = 2019
HISTORICAL_URL = (
    "https://storage.googleapis.com/bln-data-public/warn-layoffs/tx_historical.xlsx"
)

XLSX_MAGIC = b"PK\x03\x04"
WAF_CHALLENGE_MARKER = b"awsWafCookie"


def scrape(
    data_dir: Path = utils.WARN_DATA_DIR,
    cache_dir: Path = utils.WARN_CACHE_DIR,
) -> Path:
    cache = Cache(cache_dir)
    session = requests.Session()

    hrefs = _discover_links(_plain_fetch(session, cache))
    if not hrefs:
        logger.warning(
            "TX index page yielded no spreadsheet links (AWS WAF challenge "
            "likely); retrying via headless Chrome."
        )
        hrefs = _discover_links(_browser_fetch(session, cache))

    if hrefs:
        candidates = [(_get_year(h), [f"{ASSET_ROOT}{h}"]) for h in hrefs]
    else:
        # Last resort: the yearly URLs are predictable.
        logger.warning("TX: no links even via browser; probing constructed URLs.")
        candidates = [
            (
                year,
                [
                    f"{ASSET_ROOT}/sites/default/files/oei/docs/warn-act-listings-{year}-twc.xlsx",
                    f"{ASSET_ROOT}/sites/default/files/oei/docs/warn-act-listings-{year}.xlsx",
                ],
            )
            for year in range(FIRST_YEAR, date.today().year + 1)
        ]

    row_list: list[list] = []
    workbooks = 0
    for year, urls in candidates:
        excel_path = _download_year(session, cache_dir, year, urls)
        if excel_path is None:
            continue
        worksheet = load_workbook(filename=excel_path).worksheets[0]
        for irow, row in enumerate(worksheet.rows):
            if workbooks > 0 and irow == 0:
                continue  # keep the header only from the first workbook
            cell_list = [cell.value for cell in row]
            if cell_list[0] is None:
                continue
            row_list.append(cell_list)
        workbooks += 1

    if workbooks == 0:
        raise Exception("TX: no yearly workbooks retrievable (WAF-blocked?).")

    # Workbooks vary in trailing empty columns; strip them from the header so
    # the emitted CSV header doesn't depend on which year came first.
    header = row_list[0]
    while header and header[-1] in (None, ""):
        header.pop()

    # Historical data (pre-2019) from BLN's archived workbook, trimmed to the
    # same columns as the yearly files — unchanged from upstream.
    excel_path = cache.download("tx/historical.xlsx", HISTORICAL_URL)
    worksheet = load_workbook(filename=excel_path).worksheets[0]
    for i, row in enumerate(worksheet.rows):
        if i == 0:
            continue
        select_columns = [
            row[8],  # NOTICE_DATE
            row[0],  # JOB_SITE_NAME
            row[2],  # COUNTY_NAME
            row[5],  # WDA_NAME
            row[6],  # TOTAL_LAYOFF_NUMBER
            row[7],  # LayOff_Date
            row[11],  # WFDD_RECEIVED_DATE
            row[1],  # CITY_NAME
        ]
        row_list.append([c.value for c in select_columns])

    data_path = data_dir / "tx.csv"
    utils.write_rows_to_csv(data_path, row_list)
    return data_path


def _plain_fetch(session: requests.Session, cache: Cache) -> str:
    """Fetch the index page with plain HTTP; returns '' on failure."""
    try:
        page = session.get(INDEX_URL, timeout=60)
        logger.debug("TX index page status %s", page.status_code)
    except Exception as exc:
        logger.warning("TX index page fetch failed: %s", exc)
        return ""
    cache.write("tx/source.html", page.text)
    return page.text


def _browser_fetch(session: requests.Session, cache: Cache) -> str:
    """Load the index in headless Chrome to pass the AWS WAF challenge.

    Returns the rendered HTML and, as a side effect, copies the WAF cookies
    and matching user agent into ``session`` so asset downloads pass too.
    Returns '' if selenium/Chrome is unavailable or the challenge doesn't
    clear.
    """
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options as ChromeOptions
    except ImportError:
        logger.warning("TX: selenium not installed; cannot run browser fallback.")
        return ""

    options = ChromeOptions()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    try:
        driver = webdriver.Chrome(options=options)
    except Exception as exc:
        logger.warning("TX: could not start Chrome (%s).", exc)
        return ""

    try:
        driver.set_page_load_timeout(60)
        driver.get(INDEX_URL)
        # AWS WAF solves its challenge and reloads; poll for real content.
        deadline = time() + 45
        html = driver.page_source
        while "warn-act-listings-" not in html and time() < deadline:
            sleep(2)
            html = driver.page_source
        cache.write("tx/source.html", html)
        for c in driver.get_cookies():
            session.cookies.set(c["name"], c["value"], domain=c.get("domain"))
        ua = driver.execute_script("return navigator.userAgent")
        session.headers["User-Agent"] = ua  # WAF token is fingerprint-bound
        if "warn-act-listings-" not in html:
            logger.warning("TX: WAF challenge did not clear within 45s.")
            return ""
        return html
    finally:
        driver.quit()


def _discover_links(html: str) -> list[str]:
    """Extract yearly-workbook hrefs from index HTML; [] when none."""
    if not html:
        return []
    soup = BeautifulSoup(html, "html5lib")
    hrefs = [a.get("href") for a in soup.find_all("a", href=HREF_PATTERN)]
    return [h for h in hrefs if _get_year(h) >= FIRST_YEAR]


def _get_year(url: str) -> int:
    """Plucks the year from a workbook URL (upstream logic)."""
    m = re.match(r".*-(\d{4})(.*)$", url, re.I)
    assert m is not None
    return int(m.group(1)[-4:])


def _download_year(
    session: requests.Session, cache_dir: Path, year: int, urls: list[str]
) -> Path | None:
    """Fetch the first URL that returns a genuine xlsx; None if none do."""
    for url in urls:
        try:
            r = session.get(url, timeout=60)
        except Exception as exc:
            logger.warning("TX %s: fetch failed (%s)", year, exc)
            continue
        sleep(random.uniform(2, 4))
        if r.status_code != 200 or not r.content.startswith(XLSX_MAGIC):
            challenged = WAF_CHALLENGE_MARKER in (r.content or b"")[:2000]
            logger.info(
                "TX %s: %s -> status %s%s, not a workbook",
                year, url, r.status_code, " (WAF challenge)" if challenged else "",
            )
            continue
        excel_path = cache_dir / f"tx/{year}.xlsx"
        excel_path.parent.mkdir(parents=True, exist_ok=True)
        excel_path.write_bytes(r.content)
        return excel_path
    logger.warning("TX %s: no retrievable workbook", year)
    return None

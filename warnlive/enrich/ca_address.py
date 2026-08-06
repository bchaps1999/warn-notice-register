"""Backfill California site addresses from EDD's fiscal-year WARN reports.

The processed CSV we ingest daily carries only a city for CA notices (a
raw `address` field exists but EDD only began filling it in 2026). EDD's
own fiscal-year WARN report, however, has printed a per-notice street
address since FY 2021-22 — PDFs for closed years, the twice-weekly
warn_report1.xlsx for the current one. This parses those, then fills
notices.site_address by employer + notice date (city as tiebreaker).
"""

from __future__ import annotations

import datetime as dt
import logging
import re
from dataclasses import dataclass
from pathlib import Path

import requests

logger = logging.getLogger(__name__)

_BASE = "https://edd.ca.gov/siteassets/files/jobs_and_training/warn/"
# Address column exists from FY 2021-22 on; earlier reports print city only.
REPORT_URLS = [
    _BASE + "warn-report-for-7-1-2021-to-06-30-2022.pdf",
    _BASE + "warn-report-for-7-1-2022-to-06-30-2023.pdf",
    _BASE + "warn-report-for-7-1-2023-to-06-30-2024.pdf",
    _BASE + "warn-report-for-7-1-2024-to-06-30-2025.pdf",
    _BASE + "warn-report-for-7-1-25-to-6-30-26.pdf",
    _BASE + "warn_report1.xlsx",  # current fiscal year, updated Tue/Thu
]

_DATE_RE = re.compile(r"(\d{1,2})/(\d{1,2})/(\d{4})")
_STREET_RE = re.compile(r"^\d{1,6}[\w./-]*\s+[A-Za-z]")
_WS = re.compile(r"\s+")


@dataclass
class CaRecord:
    company: str
    notice_date: str | None  # ISO
    effective_date: str | None
    address: str
    source_file: str


def _iso(value) -> str | None:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    m = _DATE_RE.search(str(value or ""))
    if not m:
        return None
    month, day, year = (int(g) for g in m.groups())
    try:
        return dt.date(year, month, day).isoformat()
    except ValueError:
        return None


def _clean(addr) -> str | None:
    text = _WS.sub(" ", str(addr or "")).strip(" ,;")
    return text if _STREET_RE.search(text) else None


def parse_pdf(path: Path) -> list[CaRecord]:
    import pdfplumber

    records = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                for row in table:
                    cells = [(_WS.sub(" ", c).strip() if c else "") for c in row]
                    if len(cells) < 5:
                        continue
                    # layout: notice, processed, effective, company, ...,
                    # address last
                    notice, effective = _iso(cells[0]), _iso(cells[2])
                    addr = _clean(cells[-1])
                    if not notice or not addr or not cells[3]:
                        continue
                    records.append(CaRecord(
                        company=cells[3], notice_date=notice,
                        effective_date=effective, address=addr,
                        source_file=path.name,
                    ))
    return records


def parse_xlsx(path: Path) -> list[CaRecord]:
    import warnings

    import openpyxl

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = next((s for s in wb.sheetnames if s.strip() == "Detailed WARN Report"), None)
    if sheet is None:
        return []
    records = []
    idx = None
    for row in wb[sheet].iter_rows(values_only=True):
        vals = ["" if c is None else _WS.sub(" ", str(c)).strip() for c in row]
        if idx is None:
            if "Company" in vals and "Address" in vals:
                idx = {name: vals.index(name) for name in
                       ("Company", "Address", "Notice Date", "Effective Date")}
            continue
        addr = _clean(vals[idx["Address"]])
        notice = _iso(row[idx["Notice Date"]])
        if not addr or not notice or not vals[idx["Company"]]:
            continue
        records.append(CaRecord(
            company=vals[idx["Company"]], notice_date=notice,
            effective_date=_iso(row[idx["Effective Date"]]), address=addr,
            source_file=path.name,
        ))
    return records


def collect_records(cache_dir: Path) -> list[CaRecord]:
    cache_dir.mkdir(parents=True, exist_ok=True)
    session = requests.Session()
    session.headers["User-Agent"] = "warn-live/1.0"
    records = []
    for url in REPORT_URLS:
        name = url.rsplit("/", 1)[-1]
        path = cache_dir / name
        # the current-FY workbook changes twice a week; always refetch it
        if not path.exists() or name == "warn_report1.xlsx":
            resp = session.get(url, timeout=180)
            resp.raise_for_status()
            path.write_bytes(resp.content)
        try:
            parsed = parse_xlsx(path) if path.suffix == ".xlsx" else parse_pdf(path)
        except Exception as exc:  # noqa: BLE001 — one bad report shouldn't sink the run
            logger.warning("ca-address: %s failed: %s", name, exc)
            continue
        logger.info("ca-address: %s -> %d records", name, len(parsed))
        records.extend(parsed)
    return records
